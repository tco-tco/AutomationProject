import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

        # font = Font(name='Calibri',
        #          size=11,
        #          bold=False,
        #          italic=False,
        #          vertAlign=None,
        #          underline='none',
        #          strike=False,
        #          color='FF000000')

class ExcelDosya():
    def __init__(self,satir,sutun,deger):
        self.workbook=openpyxl.load_workbook("C:/Users/tugbacanan.oguz/Desktop/CED_TCO.xlsx")
        self.satir=satir
        self.sutun=sutun
        self.deger=deger

    def sıraYaz(self):           

        sheet1=self.workbook["Sayfa1"]
        sheet1.cell(row=self.satir,column=self.sutun, value=self.deger)
        sheet1.cell(row=self.satir,column=self.sutun, value=self.deger).font=Font(bold=True)
        thin = Side(border_style="thin", color="000000")
        sheet1.cell(row=self.satir,column=self.sutun, value=self.deger).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        
        self.workbook.save("C:/Users/tugbacanan.oguz/Desktop/CED_TCO.xlsx")

    def ililcetarihYaz(self):           

        sheet1=self.workbook["Sayfa1"]
        sheet1.cell(row=self.satir,column=self.sutun, value=self.deger)
        thin = Side(border_style="thin", color="000000")
        sheet1.cell(row=self.satir,column=self.sutun, value=self.deger).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        #sheet1.cell(row=self.satir,column=self.sutun, value=self.deger).font=Font(bold=True)
        
        self.workbook.save("C:/Users/tugbacanan.oguz/Desktop/CED_TCO.xlsx")

    def xYaz(self):           

        sheet1=self.workbook["Sayfa1"]
        sheet1.cell(row=self.satir,column=self.sutun, value=self.deger)
        # sheet1.cell(row=self.satir,column=self.sutun, value=self.deger).fill=PatternFill("solid",fgColor="00FF8080")
        thin = Side(border_style="thin", color="000000")
        sheet1.cell(row=self.satir,column=self.sutun, value=self.deger).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        sheet1.cell(row=self.satir,column=self.sutun, value=self.deger).alignment = Alignment(horizontal="center", vertical="center")        

        self.workbook.save("C:/Users/tugbacanan.oguz/Desktop/CED_TCO.xlsx")

    def xSil(self):           

        sheet1=self.workbook["Sayfa1"]
        sheet1.cell(row=self.satir,column=self.sutun, value=self.deger)
        
        thin = Side(border_style="thin", color="000000")
        sheet1.cell(row=self.satir,column=self.sutun, value=self.deger).border = Border(top=thin, left=thin, right=thin, bottom=thin)
              

        self.workbook.save("C:/Users/tugbacanan.oguz/Desktop/CED_TCO.xlsx")

    def fillYellow(self):
        sheet1=self.workbook["Sayfa1"]
        sheet1.cell(row=self.satir,column=self.sutun, value=self.deger).fill=PatternFill("solid",fgColor="00FFFF00")
        thin = Side(border_style="thin", color="000000")
        sheet1.cell(row=self.satir,column=self.sutun, value=self.deger).border = Border(top=thin, left=thin, right=thin, bottom=thin)
              

        self.workbook.save("C:/Users/tugbacanan.oguz/Desktop/CED_TCO.xlsx")

    def oOkYaz(self):           

        sheet1=self.workbook["Sayfa1"]
        sheet1.cell(row=self.satir,column=self.sutun, value=self.deger)
        # sheet1.cell(row=self.satir,column=self.sutun, value=self.deger).fill=PatternFill("solid",fgColor="00FF8080")
        thin = Side(border_style="thin", color="000000")
        sheet1.cell(row=self.satir,column=self.sutun, value=self.deger).border = Border(top=thin, left=thin, right=thin, bottom=thin)       

        self.workbook.save("C:/Users/tugbacanan.oguz/Desktop/CED_TCO.xlsx")

    





